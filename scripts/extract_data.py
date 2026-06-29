#!/usr/bin/env python3
"""Convert Rack_Fender_Kickstand_Compatibility_202507update.xlsx into data/data.json + data/images/."""
import json
import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def sanitize_sku(raw_sku):
    text = str(raw_sku).strip()
    text = re.sub(r"[^A-Za-z0-9]+", "-", text)
    return text.strip("-")


def parse_compat_cell(raw):
    if raw is None:
        raise ValueError("compatibility cell is empty")
    text = str(raw).strip()
    if text == "\xad":
        return {"status": "accessory"}
    if text == "X":
        return {"status": "no"}
    if text.startswith("○"):
        rest = text[1:].strip()
        if not rest:
            return {"status": "yes"}
        note = rest.lstrip("*").strip()
        note = note.replace("moutning", "mounting")
        return {"status": "yes", "note": note}
    raise ValueError(f"unrecognized compatibility symbol: {raw!r}")


ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Rack_Fender_Kickstand_Compatibility_202507update.xlsx"
DATA_DIR = ROOT / "data"
IMAGES_DIR = DATA_DIR / "images"

DRAWING_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}
REL_EMBED_ATTR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"


def find_group_start(ws, label, header_row=2, max_col=80):
    for c in range(1, max_col):
        if ws.cell(header_row, c).value == label:
            return c
    raise ValueError(f"could not find column group {label!r}")


def find_group_end(ws, start_col, name_row=3, max_col=80):
    c = start_col
    while c <= max_col and ws.cell(name_row, c).value is not None:
        c += 1
    return c


def extract_models(ws):
    """Returns (models_for_json, [(row, model) for compat lookup])."""
    models = []
    model_rows = []
    current_brand = None
    for r in range(1, ws.max_row + 1):
        brand_cell = ws.cell(r, 1).value
        name_cell = ws.cell(r, 2).value
        if brand_cell in ("GIANT", "Liv"):
            current_brand = brand_cell
            continue
        if current_brand is None or not name_cell:
            continue
        name = str(name_cell).strip()
        model = {"id": slugify(f"{current_brand}-{name}"), "brand": current_brand, "name": name}
        models.append(model)
        model_rows.append((r, model))
    return models, model_rows


def load_picture_anchors(zf, drawing_path, rels_path):
    rels = {el.get("Id"): el.get("Target") for el in ET.fromstring(zf.read(rels_path))}
    anchors = []
    for anchor in ET.fromstring(zf.read(drawing_path)):
        frm = anchor.find("xdr:from", DRAWING_NS)
        if frm is None:
            continue
        col = int(frm.find("xdr:col", DRAWING_NS).text)
        row = int(frm.find("xdr:row", DRAWING_NS).text)
        blip = anchor.find(".//a:blip", DRAWING_NS)
        if blip is None:
            continue
        rid = blip.get(REL_EMBED_ATTR)
        target = rels.get(rid)
        if target:
            anchors.append((col, row, "xl/media/" + target.split("/")[-1]))
    return anchors


def save_image(zf, media_zip_path, dest_dir, base_name):
    dest_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(media_zip_path).suffix
    dest = dest_dir / f"{base_name}{ext}"
    dest.write_bytes(zf.read(media_zip_path))
    return dest


def extract_products(ws, zf, anchors, col_range, category):
    picture_row0 = 4  # row 5 (1-indexed) holds product photos
    col_to_media = {col + 1: media for col, row, media in anchors if row == picture_row0}
    products = []
    col_to_id = {}
    dest_dir = IMAGES_DIR / category
    for col in col_range:
        name = ws.cell(3, col).value
        raw_sku = ws.cell(4, col).value
        if name is None or raw_sku is None:
            continue
        name = " ".join(str(name).split())
        pid = sanitize_sku(raw_sku)
        image_rel = None
        media = col_to_media.get(col)
        if media:
            dest = save_image(zf, media, dest_dir, pid)
            image_rel = f"images/{category}/{dest.name}"
        products.append({"id": pid, "sku": str(raw_sku).strip(), "name": name, "image": image_rel})
        col_to_id[col] = pid
    return products, col_to_id


def extract_compat(ws, model_rows, *col_to_id_maps):
    col_to_id = {}
    for m in col_to_id_maps:
        col_to_id.update(m)
    compat = {}
    for row, model in model_rows:
        entry = {}
        for col, pid in col_to_id.items():
            entry[pid] = parse_compat_cell(ws.cell(row, col).value)
        compat[model["id"]] = entry
    return compat


def extract_kickstands(ws, zf, anchors):
    row_to_media = {row + 1: media for _col, row, media in anchors}
    dest_dir = IMAGES_DIR / "kickstand"
    products = []
    for r in range(3, ws.max_row + 1):
        raw_sku = ws.cell(r, 1).value
        if raw_sku is None:
            continue
        sku = str(raw_sku).strip()
        name = str(ws.cell(r, 4).value or "").strip()
        bikes_text = str(ws.cell(r, 5).value or "").strip()
        ebikes_text = str(ws.cell(r, 6).value or "").strip()
        image_rel = None
        media = row_to_media.get(r)
        if media:
            dest = save_image(zf, media, dest_dir, sku)
            image_rel = f"images/kickstand/{dest.name}"
        products.append(
            {
                "id": sku,
                "sku": sku,
                "name": name,
                "image": image_rel,
                "compatibleBikesText": bikes_text,
                "compatibleEbikesText": ebikes_text,
            }
        )
    return products


def build_data(xlsx_path=XLSX_PATH):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_fr = wb["Fender_Rack"]
    ws_ks = wb["Kickstand"]

    fender_start = find_group_start(ws_fr, "Fender")
    rack_start = find_group_start(ws_fr, "Rack")
    rack_end = find_group_end(ws_fr, rack_start)

    models, model_rows = extract_models(ws_fr)

    shutil.rmtree(IMAGES_DIR, ignore_errors=True)

    with zipfile.ZipFile(xlsx_path) as zf:
        anchors_fr = load_picture_anchors(zf, "xl/drawings/drawing1.xml", "xl/drawings/_rels/drawing1.xml.rels")
        anchors_ks = load_picture_anchors(zf, "xl/drawings/drawing2.xml", "xl/drawings/_rels/drawing2.xml.rels")

        fenders, fender_col_to_id = extract_products(
            ws_fr, zf, anchors_fr, range(fender_start, rack_start), "fender"
        )
        racks, rack_col_to_id = extract_products(ws_fr, zf, anchors_fr, range(rack_start, rack_end), "rack")
        kickstands = extract_kickstands(ws_ks, zf, anchors_ks)

    compat = extract_compat(ws_fr, model_rows, fender_col_to_id, rack_col_to_id)

    return {
        "legend": {"yes": "適用", "no": "不適用", "accessory": "僅限配件或特定情況"},
        "models": models,
        "fenders": fenders,
        "racks": racks,
        "kickstands": kickstands,
        "compat": compat,
    }


def main():
    data = build_data()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(DATA_DIR / "data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(
        f"models={len(data['models'])} fenders={len(data['fenders'])} "
        f"racks={len(data['racks'])} kickstands={len(data['kickstands'])}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
